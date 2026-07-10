#!/usr/bin/env python3
from pathlib import Path
import openpyxl

def main():
    base = Path(__file__).resolve().parents[2]
    p = base / 'doc' / 'Fixture-Copa-Mundial-FIFA-2026_ClasesExcel.xlsx'
    if not p.exists():
        print('NO_ENCONTRADO', p)
        return
    wb = openpyxl.load_workbook(p, read_only=True)
    print('Hojas:', wb.sheetnames)
    for s in wb.sheetnames:
        print('\n--- Hoja:', s)
        ws = wb[s]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10:
                break
            rows.append(list(row))
        for r in rows:
            print(r)

if __name__ == '__main__':
    main()
